import os
import openpyxl
from openpyxl import Workbook
import datetime

EXCEL_FILE = "registro_asistente.xlsx"

def log_to_excel(sender, is_group, message, action_taken, details):
    """Guarda toda la información procesada en un Excel local."""
    try:
        # Si el archivo no existe, crearlo con las cabeceras
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Registros"
            ws.append(["Fecha", "Hora", "Remitente", "Es Grupo?", "Mensaje del Usuario", "Acción de la IA", "Detalles / Respuesta"])
            wb.save(EXCEL_FILE)
        
        # Cargar el archivo y añadir la nueva fila
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        now = datetime.datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")
        grupo_str = "Sí" if is_group else "No"
        
        ws.append([date_str, time_str, sender, grupo_str, message, action_taken, details])
        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"Error guardando en el archivo Excel: {e}")

def get_recent_history(limit=15):
    """Devuelve los últimos N mensajes registrados en el Excel para dar contexto a la IA."""
    try:
        if not os.path.exists(EXCEL_FILE):
            return ""
        
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        ws = wb.active
        
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows: return ""
            
        recent_rows = rows[-limit:]
        history = []
        for row in recent_rows:
            if len(row) >= 5 and row[4]: # Asegurar que hay mensaje
                sender = str(row[2]).split('@')[0] if row[2] else "Usuario"
                msg = str(row[4])
                history.append(f"{sender}: {msg}")
                
        return "\n".join(history)
    except Exception as e:
        print(f"Error leyendo historial: {e}")
        return ""
